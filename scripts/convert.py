import openpyxl  
import json  
import os  
from collections import defaultdict  
  
file_path = r"../raw/POS RAW.xlsx"  
output_dir = "../public/data"  
os.makedirs(output_dir, exist_ok=True)  
  
print("파일 로딩 중...")  
wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)  
  
REGION_SHEETS = [  
    "서울", "강원", "경기", "인천", "경남", "경북",  
    "광주", "대구", "대전", "부산", "세종", "울산",  
    "전남", "전북", "제주", "충남", "충북"  
]  
  
BLOCK_TO_SHEET = {  
    0: "서울", 1: "강원", 2: "경기", 3: "경남", 4: "경북",  
    5: "광주", 6: "대구", 7: "대전", 8: "부산", 9: "세종",  
    10: "울산", 11: "인천", 12: "전남", 13: "전북", 14: "제주",  
    15: "충남", 16: "충북"  
}  
  
SU_TO_SHEET = {  
    "중부": ["서울", "강원", "경기", "인천"],  
    "남부": ["경남", "경북", "광주", "대구", "대전", "부산",  
             "세종", "울산", "전남", "전북", "제주", "충남", "충북"]  
}  
  
SHEET_TO_SU = {  
    "서울": "중부", "강원": "중부", "경기": "중부", "인천": "중부",  
    "경남": "남부", "경북": "남부", "광주": "남부", "대구": "남부",  
    "대전": "남부", "부산": "남부", "세종": "남부", "울산": "남부",  
    "전남": "남부", "전북": "남부", "제주": "남부", "충남": "남부", "충북": "남부"  
}  
  
regions_data = {}  
cj_msref_by_region_cat = {}  
  
for region in REGION_SHEETS:  
    ws = wb[region]  
    region_cats = []  
    cj_msref_by_region_cat[region] = {}  
    for row in ws.iter_rows(min_row=7, max_row=ws.max_row, min_col=12, max_col=19, values_only=True):  
        cat = row[0]  
        if cat is None:  
            continue  
        region_cats.append({  
            "cat": cat, "cj": row[1], "competitor": row[2], "ms": row[3],  
            "cj_ref": row[4], "competitor_ref": row[5], "ms_ref": row[6], "ms_diff": row[7]  
        })  
        if row[6] is not None:  
            cj_msref_by_region_cat[region][cat] = row[6]  
    regions_data[region] = region_cats  
    print(f"  {region} 완료: {len(region_cats)}개 CAT")  
  
with open(f"{output_dir}/regions.json", "w", encoding="utf-8") as f:  
    json.dump(regions_data, f, ensure_ascii=False, separators=(',', ':'), default=str)  
print("regions.json 저장 완료!")  
  
ws_raw1 = wb["raw1"]  
region_map = {}  
for row in ws_raw1.iter_rows(min_row=3, max_row=ws_raw1.max_row, min_col=6, max_col=8, values_only=True):  
    jijeom = row[0]  
    region2 = row[2]  
    if jijeom is None or region2 is None:  
        continue  
    if jijeom not in region_map:  
        region_map[jijeom] = set()  
    region_map[jijeom].add(region2)  
  
region_map_list = {k: list(v) for k, v in region_map.items()}  
with open(f"{output_dir}/region_map.json", "w", encoding="utf-8") as f:  
    json.dump(region_map_list, f, ensure_ascii=False, separators=(',', ':'))  
print("region_map.json 저장 완료!")  
  
print("raw2 블록 구조 확인 중...")  
ws_raw2 = wb["raw2"]  
  
region2_cols = []  
for row in ws_raw2.iter_rows(min_row=3, max_row=3, values_only=True):  
    for idx, val in enumerate(row):  
        if val == '지역2':  
            region2_cols.append(idx + 1)  
  
print(f"  블록 수: {len(region2_cols)}개")  
  
sheet_rows = defaultdict(list)  
cat2_to_cat3 = {}  
  
for block_idx, start_col in enumerate(region2_cols):  
    sheet_name = BLOCK_TO_SHEET.get(block_idx)  
    if sheet_name is None:  
        continue  
    end_col = start_col + 8  
    row_count = 0  
    for row in ws_raw2.iter_rows(  
        min_row=4, max_row=ws_raw2.max_row,  
        min_col=start_col, max_col=end_col,  
        values_only=True  
    ):  
        region2      = row[0]  
        product_code = row[1]  
        product_name = row[2]  
        pos = row[3]  
        qty = row[4]  
        maker = row[5]  
        cat1         = row[6]  
        cat2         = row[7]  
        cat3         = row[8]  
  
        if region2 is None or cat2 is None or product_name is None:  
            continue  
  
        if cat2 not in cat2_to_cat3:  
            cat2_to_cat3[cat2] = set()  
        if cat3:  
            cat2_to_cat3[cat2].add(cat3)  
  
        sheet_rows[sheet_name].append(  
            (region2, product_code, product_name, pos, qty, maker, cat1, cat2, cat3)  
        )  
        row_count += 1  
  
    print(f"  블록 {block_idx+1} -> {sheet_name}: {row_count}행")  
  
cat_hierarchy = {}  
for cat2, cat3s in cat2_to_cat3.items():  
    children = sorted(list(cat3s))  
    cat_hierarchy[cat2] = {"level": "cat2", "children": children}  
    for cat3 in cat3s:  
        if cat3 not in cat_hierarchy:  
            cat_hierarchy[cat3] = {"level": "cat3", "parent": cat2, "children": []}  
  
with open(f"{output_dir}/cat_hierarchy.json", "w", encoding="utf-8") as f:  
    json.dump(cat_hierarchy, f, ensure_ascii=False, separators=(',', ':'))  
print("cat_hierarchy.json 저장 완료!")  
  
print("SU별 SKU pos/qty 합산 중...")  
su_cat_total = {su: defaultdict(float) for su in SU_TO_SHEET}  
su_sku_pos = {su: defaultdict(float) for su in SU_TO_SHEET}  
su_sku_qty = {su: defaultdict(float) for su in SU_TO_SHEET}  
su_maker_pos = {su: defaultdict(lambda: defaultdict(float)) for su in SU_TO_SHEET}  
  
for su, sheets in SU_TO_SHEET.items():  
    for sheet_name in sheets:  
        for row in sheet_rows[sheet_name]:  
            region2, product_code, product_name, pos, qty, maker, cat1, cat2, cat3 = row  
            pos_val = pos or 0  
            qty_val = qty or 0  
            pc = str(product_code) if product_code else ''  
  
            su_cat_total[su][cat2] += pos_val  
            if cat3:  
                su_cat_total[su][cat3] += pos_val  
  
            sku_key2 = (cat2, cat3 or '', pc, product_name, maker)  
            su_sku_pos[su][sku_key2] += pos_val  
            su_sku_qty[su][sku_key2] += qty_val  
            su_maker_pos[su][cat2][maker] += pos_val  
  
            if cat3:  
                sku_key3 = (cat3, pc, product_name, maker)  
                su_sku_pos[su][sku_key3] += pos_val  
                su_sku_qty[su][sku_key3] += qty_val  
                su_maker_pos[su][cat3][maker] += pos_val  
  
print("SU별 집계 완료!")  
  
print("제조사별 영본MS JSON 생성 중...")  
  
for sheet_name in REGION_SHEETS:  
    maker_dir = f"{output_dir}/maker_msref/{sheet_name}"  
    os.makedirs(maker_dir, exist_ok=True)  
  
    su = SHEET_TO_SU[sheet_name]  
  
    for cat_name, maker_map in su_maker_pos[su].items():  
        cat_total = su_cat_total[su].get(cat_name, 0)  
        if cat_total == 0:  
            continue  
  
        result = {}  
        official_cj_msref = cj_msref_by_region_cat.get(sheet_name, {}).get(cat_name)  
  
        if official_cj_msref is not None:  
            cj_official = float(official_cj_msref)  
            result['CJ'] = round(cj_official, 6)  
            remaining = 1.0 - cj_official  
            non_cj_total = sum(p for mk, p in maker_map.items() if mk != 'CJ')  
            for maker, pos in maker_map.items():  
                if maker != 'CJ':  
                    if non_cj_total > 0:  
                        result[maker] = round((pos / non_cj_total) * remaining, 6)  
                    else:  
                        result[maker] = 0  
        else:  
            for maker, pos in maker_map.items():  
                result[maker] = round(pos / cat_total, 6)  
  
        safe_name = cat_name.replace('/', '_').replace(' ', '_')  
        with open(f"{maker_dir}/{safe_name}.json", "w", encoding="utf-8") as f:  
            json.dump(result, f, ensure_ascii=False, separators=(',', ':'))  
  
    print(f"  {sheet_name}: 제조사 영본MS 완료")  
  
print("제조사별 영본MS JSON 생성 완료!")  
  
print("지역+카테고리별 JSON 생성 중...")  
  
for sheet_name in REGION_SHEETS:  
    cat_dir = f"{output_dir}/skus/{sheet_name}"  
    os.makedirs(cat_dir, exist_ok=True)  
  
    su = SHEET_TO_SU[sheet_name]  
    cat_data = defaultdict(list)  
  
    for row in sheet_rows[sheet_name]:  
        region2, product_code, product_name, pos, qty, maker, cat1, cat2, cat3 = row  
  
        pos_val = pos or 0  
        qty_val = qty or 0  
        pc = str(product_code) if product_code else ''  
  
        sku_key2 = (cat2, cat3 or '', pc, product_name, maker)  
        su_total_cat2 = su_cat_total[su].get(cat2, 0)  
        ms_ref_cat2 = su_sku_pos[su].get(sku_key2, 0) / su_total_cat2 if su_total_cat2 > 0 else 0  
        su_qty2 = su_sku_qty[su].get(sku_key2, 0)  
        su_pos2 = su_sku_pos[su].get(sku_key2, 0)  
        price_ref_cat2 = (su_pos2 / su_qty2) * 1000000 if su_qty2 > 0 else 0  
  
        ms_ref_cat3 = 0  
        price_ref_cat3 = 0  
        if cat3:  
            sku_key3 = (cat3, pc, product_name, maker)  
            su_total_cat3 = su_cat_total[su].get(cat3, 0)  
            ms_ref_cat3 = su_sku_pos[su].get(sku_key3, 0) / su_total_cat3 if su_total_cat3 > 0 else 0  
            su_qty3 = su_sku_qty[su].get(sku_key3, 0)  
            su_pos3 = su_sku_pos[su].get(sku_key3, 0)  
            price_ref_cat3 = (su_pos3 / su_qty3) * 1000000 if su_qty3 > 0 else 0  
  
        sku_obj = {  
            "r2": region2, "pc": pc, "pn": product_name,  
            "pos": round(pos_val, 4), "qty": round(qty_val, 2),  
            "mk": maker, "c2": cat2, "c3": cat3,  
            "mr": round(ms_ref_cat2, 6), "mr3": round(ms_ref_cat3, 6),  
            "pr": round(price_ref_cat2, 0), "pr3": round(price_ref_cat3, 0)  
        }  
  
        cat_data[cat2].append(sku_obj)  
        if cat3:  
            cat_data[cat3].append(sku_obj)  
  
    for cat_name, skus in cat_data.items():  
        safe_name = cat_name.replace('/', '_').replace(' ', '_')  
        with open(f"{cat_dir}/{safe_name}.json", "w", encoding="utf-8") as f:  
            json.dump(skus, f, ensure_ascii=False, separators=(',', ':'), default=str)  
  
    print(f"  {sheet_name}: {len(cat_data)}개 CAT 완료")  
  
print("전체 변환 완료!")  
